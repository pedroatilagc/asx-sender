import re
from openpyxl import load_workbook, Workbook
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from .models import Contato, normalizar_numero
from .serializers import ContatoSerializer


REGEX_VALIDO = re.compile(r'^55\d{10,11}$')


class ContatoViewSet(viewsets.ModelViewSet):
    queryset         = Contato.objects.all()
    serializer_class = ContatoSerializer
    parser_classes   = [JSONParser, MultiPartParser, FormParser]

    def get_queryset(self):
        qs      = Contato.objects.all()
        opt_out = self.request.query_params.get('opt_out')
        busca   = self.request.query_params.get('busca')
        if opt_out in ('true', 'false'):
            qs = qs.filter(opt_out=(opt_out == 'true'))
        if busca:
            qs = qs.filter(numero__icontains=busca) | qs.filter(nome__icontains=busca)
        return qs

    @action(detail=False, methods=['post'], url_path='importar')
    def importar(self, request):
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            return Response({'error': 'Nenhum arquivo enviado.'}, status=400)

        try:
            wb = load_workbook(arquivo, data_only=True)
            ws = wb.active
        except Exception:
            return Response({'error': 'Arquivo XLSX inválido.'}, status=400)

        cabecalhos = [str(c.value or '').strip().lower() for c in ws[1]]
        try:
            idx_nome   = next(i for i, h in enumerate(cabecalhos) if h in ('nome', 'name'))
            idx_numero = next(i for i, h in enumerate(cabecalhos) if h in ('numero', 'número', 'number', 'telefone', 'phone'))
        except StopIteration:
            return Response({'error': 'Cabeçalhos não encontrados. Use "nome" e "numero".'}, status=400)

        criados    = 0
        ignorados  = 0
        invalidos  = 0
        existentes = set(Contato.objects.values_list('numero', flat=True))

        for linha in ws.iter_rows(min_row=2, values_only=True):
            if linha[idx_numero] is None:
                continue
            nome   = str(linha[idx_nome] or '').strip()
            numero = normalizar_numero(linha[idx_numero])
            if not REGEX_VALIDO.match(numero):
                invalidos += 1
                continue
            if numero in existentes:
                ignorados += 1
                continue
            Contato.objects.create(nome=nome, numero=numero)
            existentes.add(numero)
            criados += 1

        return Response({'criados': criados, 'ignorados': ignorados, 'invalidos': invalidos})

    @action(detail=False, methods=['get'], url_path='exportar')
    def exportar(self, request):
        opt_out = request.query_params.get('opt_out')
        qs      = Contato.objects.all()
        if opt_out in ('true', 'false'):
            qs = qs.filter(opt_out=(opt_out == 'true'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Contatos'
        ws.append(['nome', 'numero', 'opt_out', 'criado_em'])
        for c in qs:
            ws.append([c.nome, c.numero, 'sim' if c.opt_out else 'não', c.criado_em.strftime('%Y-%m-%d %H:%M')])

        resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename=contatos.xlsx'
        wb.save(resp)
        return resp

    @action(detail=False, methods=['post'], url_path='remover-duplicados')
    def remover_duplicados(self, request):
        vistos    = set()
        removidos = 0
        for c in Contato.objects.all().order_by('criado_em'):
            if c.numero in vistos:
                c.delete()
                removidos += 1
            else:
                vistos.add(c.numero)
        return Response({'removidos': removidos})

    @action(detail=False, methods=['post'], url_path='limpar-invalidos')
    def limpar_invalidos(self, request):
        removidos = 0
        for c in Contato.objects.all():
            if not REGEX_VALIDO.match(c.numero):
                c.delete()
                removidos += 1
        return Response({'removidos': removidos})

    @action(detail=False, methods=['post'], url_path='limpar-tudo')
    def limpar_tudo(self, request):
        total = Contato.objects.count()
        Contato.objects.all().delete()
        return Response({'removidos': total})
