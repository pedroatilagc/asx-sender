import re
import csv
import io
import time
import random
import threading
from datetime import datetime

from openpyxl import load_workbook
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
import requests

from .models import Campanha, ContatoCampanha
from .serializers import CampanhaSerializer, ContatoCampanhaSerializer
from .configuracoes import ler_config, salvar_config
from dispositivos.models import Dispositivo
from modelos.models import Modelo
from contatos.models import Contato

from django.conf import settings

UAZAPI_URL_CAMPANHAS = settings.UAZAPI_URL

_threads = {}
_paused  = {}
_stopped = {}


def normalizar_numero(numero):
    digitos = re.sub(r'\D', '', str(numero or ''))
    if not digitos:
        return ''
    if not digitos.startswith('55'):
        digitos = '55' + digitos
    return digitos


def enviar_mensagem_evolution_OLD(instance_name, numero, mensagem, modelo=None):
    try:
        # Mensagem com botões
        if modelo and modelo.tipo == 'botoes' and modelo.botoes:
            botoes_payload = []
            for b in modelo.botoes:
                tipo   = b.get('tipo', 'reply')
                titulo = b.get('titulo', '')
                if tipo == 'url':
                    botoes_payload.append({
                        'type':        'url',
                        'displayText': titulo,
                        'url':         b.get('url', ''),
                    })
                elif tipo == 'copiar':
                    botoes_payload.append({
                        'type':        'copy',
                        'displayText': titulo,
                        'copyCode':    b.get('codigo', ''),
                    })
                else:  # reply ou optout
                    botoes_payload.append({
                        'type':        'reply',
                        'displayText': titulo,
                        'id':          f'btn_{titulo[:20].lower().replace(" ", "_")}',
                    })

            payload = {
                'number':      numero,
                'title':       mensagem.split('\n')[0][:60] if mensagem else '',
                'description': mensagem,
                'footer':      modelo.rodape or '',
                'buttons':     botoes_payload,
            }
            res = requests.post(
                f'{EVOLUTION_URL}/message/sendButtons/{instance_name}',
                json=payload,
                headers=HEADERS,
                timeout=30,
            )
        else:
            # Texto puro
            payload = {'number': numero, 'text': mensagem}
            res = requests.post(
                f'{EVOLUTION_URL}/message/sendText/{instance_name}',
                json=payload,
                headers=HEADERS,
                timeout=30,
            )

        return res.status_code in [200, 201], res.text
    except Exception as e:
        return False, str(e)


def enviar_mensagem_uazapi(instance_token, numero, mensagem, modelo=None):
    try:
        if modelo and modelo.tipo == 'botoes' and modelo.botoes:
            choices = []
            for b in modelo.botoes:
                tipo   = b.get('tipo', '')
                titulo = b.get('titulo', '')
                if tipo == 'url':
                    choices.append(f"{titulo}|{b.get('url', '')}")
                elif tipo == 'copiar':
                    choices.append(f"{titulo}|copy:{b.get('codigo', '')}")
                else:  # optout / reply
                    choices.append(f"{titulo}|{titulo.lower().replace(' ', '_')}")

            payload = {
                'number':     numero,
                'type':       'button',
                'text':       mensagem,
                'choices':    choices,
                'footerText': modelo.rodape or 'ASX Sender',
            }
            res = requests.post(
                f'{UAZAPI_URL_CAMPANHAS}/send/menu',
                json=payload,
                headers={'token': instance_token, 'Content-Type': 'application/json'},
                timeout=30,
            )
        else:
            payload = {'number': numero, 'text': mensagem}
            res = requests.post(
                f'{UAZAPI_URL_CAMPANHAS}/send/text',
                json=payload,
                headers={'token': instance_token, 'Content-Type': 'application/json'},
                timeout=30,
            )

        return res.status_code in [200, 201], res.text
    except Exception as e:
        return False, str(e)


def processar_campanha(campanha_id):
    try:
        campanha = Campanha.objects.get(id=campanha_id)
    except Campanha.DoesNotExist:
        return

    config     = ler_config()
    instancias_qs = campanha.instancias.filter(status='open')
    instancias    = list(instancias_qs.values_list('instance_name', 'instance_token'))

    if not instancias:
        campanha.status = 'cancelada'
        campanha.save()
        return

    campanha.status      = 'enviando'
    campanha.iniciado_em = timezone.now()
    campanha.save()

    contatos                 = list(campanha.contatos.filter(status='pendente'))
    idx_inst                 = 0
    enviados_desde_suspensao = 0

    for contato in contatos:
        if _stopped.get(campanha_id):
            break

        while _paused.get(campanha_id):
            time.sleep(1)
            if _stopped.get(campanha_id):
                break

        if _stopped.get(campanha_id):
            break

        # Verificar opt-out
        if Contato.objects.filter(numero=contato.numero, opt_out=True).exists():
            contato.status = 'pulado'
            contato.erro   = 'Opt-out'
            contato.save()
            campanha.pulados += 1
            campanha.save()
            continue

        # Determinar mensagem
        if campanha.mensagem_override:
            mensagem = campanha.mensagem_override
        elif campanha.modelo:
            mensagem = campanha.modelo.mensagem
        else:
            contato.status = 'falhou'
            contato.erro   = 'Sem mensagem configurada'
            contato.save()
            campanha.falhas += 1
            campanha.save()
            continue

        # Substituir variáveis
        nome_contato = contato.nome or 'cliente'
        mensagem     = mensagem.replace('{{name}}', nome_contato).replace('{{nome}}', nome_contato)

        # Rotação round-robin de instâncias
        instance_name, instance_token = instancias[idx_inst % len(instancias)]
        idx_inst          += 1
        contato.instancia = instance_name

        modelo_obj = campanha.modelo if not campanha.mensagem_override else None
        sucesso, resposta = enviar_mensagem_uazapi(instance_token, contato.numero, mensagem, modelo=modelo_obj)

        if sucesso:
            contato.status     = 'enviado'
            contato.enviado_em = timezone.now()
            campanha.enviados += 1
            Contato.objects.get_or_create(
                numero=contato.numero,
                defaults={'nome': contato.nome}
            )
        else:
            contato.status   = 'falhou'
            contato.erro     = resposta[:200]
            campanha.falhas += 1

        contato.save()
        campanha.save()

        enviados_desde_suspensao += 1

        if config.get('modo_suspensao') and enviados_desde_suspensao >= config.get('suspensao_apos', 40):
            pausa = random.randint(config.get('suspensao_min', 900), config.get('suspensao_max', 1800))
            time.sleep(pausa)
            enviados_desde_suspensao = 0
        else:
            intervalo = random.randint(config.get('intervalo_min', 20), config.get('intervalo_max', 90))
            time.sleep(intervalo)

    if not _stopped.get(campanha_id):
        campanha.status       = 'concluida'
        campanha.concluido_em = timezone.now()
    else:
        campanha.status = 'cancelada'

    campanha.save()
    _threads.pop(campanha_id, None)
    _paused.pop(campanha_id, None)
    _stopped.pop(campanha_id, None)


class CampanhaViewSet(viewsets.ModelViewSet):
    queryset         = Campanha.objects.all()
    serializer_class = CampanhaSerializer
    parser_classes   = [JSONParser, MultiPartParser, FormParser]

    @action(detail=False, methods=['post'], url_path='importar-base')
    def importar_base(self, request):
        arquivo = request.FILES.get('arquivo')
        if not arquivo:
            return Response({'error': 'Nenhum arquivo enviado.'}, status=400)

        nome_arquivo = arquivo.name.lower()

        try:
            if nome_arquivo.endswith('.csv'):
                conteudo = arquivo.read().decode('utf-8-sig', errors='replace')
                reader   = csv.reader(io.StringIO(conteudo))
                todas    = list(reader)
                if not todas:
                    return Response({'error': 'Arquivo CSV vazio.'}, status=400)
                colunas = [str(c).strip() for c in todas[0]]
                linhas  = [[str(v).strip() for v in row] for row in todas[1:] if any(v for v in row)]
            elif nome_arquivo.endswith('.xlsx') or nome_arquivo.endswith('.xls'):
                wb      = load_workbook(arquivo, data_only=True)
                ws      = wb.active
                colunas = [str(c.value or '').strip() for c in ws[1]]
                linhas  = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    linha = [str(v or '').strip() for v in row]
                    if any(v for v in linha):
                        linhas.append(linha)
            else:
                return Response({'error': 'Formato não suportado. Use CSV ou XLSX.'}, status=400)
        except Exception as e:
            return Response({'error': f'Erro ao ler arquivo: {str(e)}'}, status=400)

        if not colunas:
            return Response({'error': 'Arquivo sem cabeçalhos.'}, status=400)

        col_nome_auto   = None
        col_numero_auto = None
        for i, c in enumerate(colunas):
            cl = c.lower()
            if cl in ('nome', 'name') and col_nome_auto is None:
                col_nome_auto = i
            if cl in ('numero', 'número', 'number', 'telefone', 'celular', 'phone', 'fone') and col_numero_auto is None:
                col_numero_auto = i

        return Response({
            'colunas':         colunas,
            'linhas':          linhas[:5],
            'total_linhas':    len(linhas),
            'col_nome_auto':   col_nome_auto,
            'col_numero_auto': col_numero_auto,
            'todas_linhas':    linhas,
        })

    @action(detail=False, methods=['post'], url_path='confirmar-base')
    def confirmar_base(self, request):
        linhas     = request.data.get('linhas', [])
        idx_numero = request.data.get('idx_numero')
        idx_nome   = request.data.get('idx_nome')

        if idx_numero is None:
            return Response({'error': 'Informe a coluna de número.'}, status=400)

        contatos  = []
        invalidos = 0

        for linha in linhas:
            if idx_numero >= len(linha):
                invalidos += 1
                continue
            numero = normalizar_numero(linha[idx_numero])
            if not re.match(r'^55\d{10,11}$', numero):
                invalidos += 1
                continue
            nome = str(linha[idx_nome]).strip() if idx_nome is not None and idx_nome < len(linha) else ''
            contatos.append({'nome': nome, 'numero': numero})

        return Response({
            'contatos': contatos,
            'total':    len(contatos),
            'invalidos': invalidos,
        })

    @action(detail=False, methods=['post'], url_path='criar-e-disparar')
    def criar_e_disparar(self, request):
        nome              = request.data.get('nome', '').strip()
        modelo_id         = request.data.get('modelo_id')

        tipo_override     = request.data.get('tipo_override', 'texto')
        instancia_ids     = request.data.get('instancia_ids', [])
        contatos_data     = request.data.get('contatos', [])
        agendado_para     = request.data.get('agendado_para')
        mensagem_override = (request.data.get('mensagem_override') or '').strip()

        if not nome:
            return Response({'error': 'Nome da campanha é obrigatório.'}, status=400)
        if not contatos_data:
            return Response({'error': 'Nenhum contato importado.'}, status=400)
        if not instancia_ids:
            return Response({'error': 'Selecione pelo menos uma instância.'}, status=400)
        if not modelo_id and not mensagem_override:
            return Response({'error': 'Selecione um modelo ou digite uma mensagem.'}, status=400)

        campanha = Campanha.objects.create(
            nome=nome,
            mensagem_override=mensagem_override,
            tipo_override=tipo_override,
            status='rascunho',
            total=len(contatos_data),
        )

        if modelo_id:
            try:
                campanha.modelo = Modelo.objects.get(id=int(modelo_id))
                campanha.save()
            except (Modelo.DoesNotExist, ValueError):
                pass

        instancias = Dispositivo.objects.filter(id__in=instancia_ids)
        campanha.instancias.set(instancias)

        for c in contatos_data:
            ContatoCampanha.objects.create(
                campanha=campanha,
                nome=c.get('nome', ''),
                numero=c.get('numero', ''),
            )

        if agendado_para:
            try:
                campanha.agendado_para = datetime.fromisoformat(agendado_para)
                campanha.status = 'rascunho'
                campanha.save()
                return Response({**CampanhaSerializer(campanha).data, 'agendada': True}, status=201)
            except Exception as e:
                campanha.delete()
                return Response({'error': f'Data inválida: {str(e)}'}, status=400)

        t = threading.Thread(target=processar_campanha, args=(campanha.id,), daemon=True)
        _threads[campanha.id] = t
        t.start()

        return Response(CampanhaSerializer(campanha).data, status=201)

    @action(detail=True, methods=['post'], url_path='pausar')
    def pausar(self, request, pk=None):
        campanha = self.get_object()
        if campanha.status == 'enviando':
            _paused[campanha.id] = True
            campanha.status = 'pausada'
            campanha.save()
        return Response(CampanhaSerializer(campanha).data)

    @action(detail=True, methods=['post'], url_path='retomar')
    def retomar(self, request, pk=None):
        campanha = self.get_object()
        if campanha.status == 'pausada':
            _paused[campanha.id] = False
            campanha.status = 'enviando'
            campanha.save()
        return Response(CampanhaSerializer(campanha).data)

    @action(detail=True, methods=['post'], url_path='cancelar')
    def cancelar(self, request, pk=None):
        campanha = self.get_object()
        _stopped[campanha.id] = True
        _paused[campanha.id]  = False
        campanha.status = 'cancelada'
        campanha.save()
        return Response(CampanhaSerializer(campanha).data)

    @action(detail=True, methods=['get'], url_path='progresso')
    def progresso(self, request, pk=None):
        campanha = self.get_object()
        return Response(CampanhaSerializer(campanha).data)

    @action(detail=True, methods=['get'], url_path='contatos')
    def listar_contatos(self, request, pk=None):
        campanha = self.get_object()
        return Response(ContatoCampanhaSerializer(campanha.contatos.all(), many=True).data)


class ConfiguracaoViewSet(viewsets.ViewSet):
    def list(self, request):
        return Response(ler_config())

    def create(self, request):
        config = ler_config()
        for k, v in request.data.items():
            if k in ('intervalo_min', 'intervalo_max', 'suspensao_apos', 'suspensao_min', 'suspensao_max'):
                config[k] = int(v)
            elif k == 'modo_suspensao':
                config[k] = bool(v)
        salvar_config(config)
        return Response(config)
